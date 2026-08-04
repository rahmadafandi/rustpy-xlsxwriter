"""What happens when a DataFrame advertises Arrow but cannot deliver it.

pandas 3 always exposes ``__arrow_c_stream__``, but calling it raises
``ImportError`` when pyarrow is not installed — a very common setup, since
pyarrow is an optional pandas dependency. That used to produce a file with
only the header row: no rows, no error, no warning.

The stubs here fail the same way without needing pyarrow to be absent, so the
regression stays covered wherever the suite runs.
"""

import datetime

import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel, write_worksheet


class _Series:
    def __init__(self, values):
        self._values = values

    def tolist(self):
        return list(self._values)

    to_list = tolist


class _Dtype:
    def __init__(self, kind):
        self.kind = kind

    def __str__(self):
        return {"i": "Int64", "f": "Float64", "O": "String", "M": "Datetime"}[self.kind]


class _BrokenArrowFrame:
    """Pandas-shaped, and its Arrow stream raises like pandas without pyarrow."""

    def __init__(self, data, kinds):
        self._data = data
        self.columns = list(data)
        self.dtypes = [_Dtype(k) for k in kinds]

    def __arrow_c_stream__(self, requested_schema=None):
        raise ImportError("`Import pyarrow` failed.  Use pip or conda to install")

    def __len__(self):
        return len(next(iter(self._data.values())))

    def __getitem__(self, key):
        return _Series(self._data[key])


class _BrokenArrowPolarsFrame(_BrokenArrowFrame):
    """Same, but with the Polars accessor so the other branch is exercised."""

    columns: list

    def get_column(self, name):
        return _Series(self._data[name])

    def __getitem__(self, key):  # pragma: no cover - must not be reached
        raise AssertionError("polars frames must go through get_column")


DATA = {
    "name": ["a", "b", "c"],
    "score": [1.5, 2.5, 3.5],
    "count": [1, 2, 3],
}
KINDS = ["O", "f", "i"]


@pytest.mark.parametrize(
    "frame_cls", [_BrokenArrowFrame, _BrokenArrowPolarsFrame], ids=["pandas", "polars"]
)
def test_rows_are_written_when_the_arrow_stream_fails(tmp_path, frame_cls):
    """The regression: this used to write the header and drop every row."""
    path = tmp_path / "fallback.xlsx"
    write_worksheet(frame_cls(DATA, KINDS), str(path))

    ws = openpyxl.load_workbook(path).active
    assert [c.value for c in ws[1]] == ["name", "score", "count"]
    assert [c.value for c in ws[2]] == ["a", 1.5, 1]
    assert [c.value for c in ws[4]] == ["c", 3.5, 3]
    assert ws.max_row == 4


def test_fallback_honours_formatting_options(tmp_path):
    """It goes through the normal writer, so the options still apply."""
    path = tmp_path / "styled.xlsx"
    write_worksheet(
        _BrokenArrowFrame(DATA, KINDS),
        str(path),
        float_format="0.00",
        banded_rows="#F2F2F2",
        autofilter=True,
        totals_row={"count": "sum"},
    )

    ws = openpyxl.load_workbook(path).active
    assert ws.cell(row=2, column=2).number_format == "0.00"
    assert ws.cell(row=3, column=1).fill.fgColor.rgb == "FFF2F2F2"
    assert ws.auto_filter.ref == "A1:C4"
    assert ws["C5"].value == "=SUM(C2:C4)"


def test_fallback_handles_datetimes(tmp_path):
    path = tmp_path / "dates.xlsx"
    frame = _BrokenArrowFrame(
        {"when": [datetime.datetime(2026, 1, d) for d in (1, 2)]}, ["M"]
    )
    write_worksheet(frame, str(path))

    ws = openpyxl.load_workbook(path).active
    assert ws["A2"].value == datetime.datetime(2026, 1, 1)
    assert ws["A3"].value == datetime.datetime(2026, 1, 2)


def test_fallback_through_the_builder(tmp_path):
    path = tmp_path / "builder.xlsx"
    FastExcel(str(path)).sheet("S", _BrokenArrowFrame(DATA, KINDS)).save()
    assert openpyxl.load_workbook(path)["S"].max_row == 4


class _NotAFrame:
    """Advertises the stream but is not tabular — nothing to fall back to."""

    def __arrow_c_stream__(self, requested_schema=None):
        raise ImportError("no pyarrow")


def test_unusable_object_raises_the_original_error(tmp_path):
    """Better a clear failure than a silently empty file."""
    with pytest.raises(Exception, match="pyarrow"):
        write_worksheet(_NotAFrame(), str(tmp_path / "bad.xlsx"))


def test_real_dataframes_still_take_the_arrow_path(tmp_path):
    """The fallback must not shadow the fast path when Arrow works."""
    pd = pytest.importorskip("pandas")
    pytest.importorskip("pyarrow")
    path = tmp_path / "arrow.xlsx"
    write_worksheet(pd.DataFrame(DATA), str(path))

    ws = openpyxl.load_workbook(path).active
    assert ws.max_row == 4
    assert [c.value for c in ws[2]] == ["a", 1.5, 1]


# --- input that cannot be iterated at all ---------------------------------
# Same failure shape as the Arrow one: the loop used to be skipped on error,
# leaving an empty file and a successful return.


class _NotIterable:
    pass


@pytest.mark.parametrize(
    "data", [42, None, _NotIterable(), 3.5], ids=["int", "none", "object", "float"]
)
def test_non_iterable_input_raises(tmp_path, data):
    from rustpy_xlsxwriter import write_csv

    for writer, suffix in ((write_worksheet, ".xlsx"), (write_csv, ".csv")):
        path = tmp_path / f"bad{suffix}"
        with pytest.raises(TypeError, match="must be an iterable of dicts"):
            writer(data, str(path))


def test_iterable_of_non_dicts_still_names_the_real_problem(tmp_path):
    """A set is iterable, so it must fail on the items, not on iteration."""
    with pytest.raises(TypeError, match="must be dictionaries"):
        write_worksheet({1, 2, 3}, str(tmp_path / "bad.xlsx"))


def test_empty_input_is_still_valid(tmp_path):
    """An empty list is not an error — it just has no rows."""
    path = tmp_path / "empty.xlsx"
    write_worksheet([], str(path))
    assert openpyxl.load_workbook(path).active.max_row == 1


def test_generator_input_is_still_accepted(tmp_path):
    path = tmp_path / "gen.xlsx"
    write_worksheet((r for r in [{"a": 1}, {"a": 2}]), str(path))
    assert openpyxl.load_workbook(path).active.max_row == 3
