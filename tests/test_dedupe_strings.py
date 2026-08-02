"""Shared-string deduplication (``dedupe_strings=``)."""

import zipfile

import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel, write_worksheet, write_worksheets


def _repetitive_records(n=2000):
    statuses = ["pending", "shipped", "delivered", "cancelled"]
    return [
        {"id": i, "status": statuses[i % len(statuses)], "note": "no remarks"}
        for i in range(n)
    ]


def _shared_strings(path):
    """Return the raw xl/sharedStrings.xml, or None when absent."""
    with zipfile.ZipFile(path) as z:
        if "xl/sharedStrings.xml" not in z.namelist():
            return None
        return z.read("xl/sharedStrings.xml").decode("utf-8")


def test_off_by_default_writes_inline_strings(tmp_path):
    path = tmp_path / "inline.xlsx"
    write_worksheet(_repetitive_records(), str(path))
    assert _shared_strings(path) is None


def test_dedupe_builds_shared_string_table(tmp_path):
    path = tmp_path / "shared.xlsx"
    write_worksheet(_repetitive_records(), str(path), dedupe_strings=True)

    xml = _shared_strings(path)
    assert xml is not None
    # 4 statuses + 1 note + 3 headers, each stored exactly once.
    assert 'uniqueCount="8"' in xml
    assert xml.count("<t>delivered</t>") == 1


def _uncompressed_size(path):
    with zipfile.ZipFile(path) as z:
        return sum(i.file_size for i in z.infolist())


def test_dedupe_shrinks_repeated_text(tmp_path):
    """Long repeated strings are the case dedup is actually for.

    Note the win is on the *uncompressed* XML; ``.xlsx`` is a zip and deflate
    already collapses repetition, so the on-disk saving is far smaller — and
    for short repeated strings the shared-string table costs more than it
    saves. See ``test_short_repeats_do_not_shrink_on_disk``.
    """
    inline = tmp_path / "inline.xlsx"
    shared = tmp_path / "shared.xlsx"
    long_values = [f"{c}-{'x' * 120}" for c in "abcde"]
    records = [{"id": i, "s": long_values[i % 5]} for i in range(5000)]

    write_worksheet(records, str(inline))
    write_worksheet(records, str(shared), dedupe_strings=True)

    assert _uncompressed_size(shared) < _uncompressed_size(inline) / 2
    assert shared.stat().st_size < inline.stat().st_size


def test_short_repeats_do_not_shrink_on_disk(tmp_path):
    """Guards the documented caveat: dedup is not a free win."""
    inline = tmp_path / "inline.xlsx"
    shared = tmp_path / "shared.xlsx"
    records = _repetitive_records()

    write_worksheet(records, str(inline))
    write_worksheet(records, str(shared), dedupe_strings=True)

    # Uncompressed XML still shrinks...
    assert _uncompressed_size(shared) < _uncompressed_size(inline)
    # ...but after zip compression the table is pure overhead here.
    assert shared.stat().st_size >= inline.stat().st_size


def test_dedupe_preserves_cell_values(tmp_path):
    path = tmp_path / "values.xlsx"
    records = _repetitive_records(10)
    write_worksheet(records, str(path), sheet_name="Data", dedupe_strings=True)

    ws = openpyxl.load_workbook(path)["Data"]
    assert [c.value for c in ws[1]] == ["id", "status", "note"]
    for row_idx, record in enumerate(records, start=2):
        assert ws.cell(row=row_idx, column=1).value == record["id"]
        assert ws.cell(row=row_idx, column=2).value == record["status"]
        assert ws.cell(row=row_idx, column=3).value == record["note"]


@pytest.mark.parametrize("key", ["Repeats", "general"])
def test_multi_sheet_dedupe_is_per_sheet(tmp_path, key):
    path = tmp_path / "multi.xlsx"
    write_worksheets(
        [("Repeats", _repetitive_records(200)), ("Other", [{"x": "unique"}])],
        str(path),
        dedupe_strings={key: True},
    )

    xml = _shared_strings(path)
    assert xml is not None
    assert "<t>delivered</t>" in xml
    # "unique" only lands in the table when dedup applies to every sheet.
    assert ("<t>unique</t>" in xml) == (key == "general")


def test_fastexcel_sheet_flag(tmp_path):
    path = tmp_path / "builder.xlsx"
    (
        FastExcel(str(path))
        .sheet("Data", _repetitive_records(200), dedupe_strings=True)
        .save()
    )
    assert _shared_strings(path) is not None


def test_fastexcel_default_stays_constant_memory(tmp_path):
    path = tmp_path / "builder-default.xlsx"
    FastExcel(str(path)).sheet("Data", _repetitive_records(200)).save()
    assert _shared_strings(path) is None
