import io

import openpyxl
import pandas as pd
import pytest

from rustpy_xlsxwriter import FastExcel

from conftest import XLSX_MAGIC


class TestDataFrameSingleSheet:
    def test_content_roundtrip(self, tmp_path):
        df = pd.DataFrame(
            {"Name": ["Alice", "Bob"], "Age": [30, 25], "Score": [95.5, 88.0]}
        )
        path = str(tmp_path / "df.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert [ws.cell(1, c).value for c in range(1, 4)] == ["Name", "Age", "Score"]
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == pytest.approx(30.0)
        assert ws.cell(2, 3).value == pytest.approx(95.5)
        assert ws.cell(3, 1).value == "Bob"
        wb.close()

    def test_empty_dataframe(self, tmp_path):
        df = pd.DataFrame({"A": [], "B": []})
        path = str(tmp_path / "df_empty.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).value == "A"
        assert ws.cell(1, 2).value == "B"
        assert ws.cell(2, 1).value is None
        wb.close()

    def test_mixed_column_types(self, tmp_path):
        df = pd.DataFrame(
            {
                "int_col": [1, 2],
                "float_col": [1.5, 2.5],
                "str_col": ["a", "b"],
            }
        )
        path = str(tmp_path / "df_mixed.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == pytest.approx(1.5)
        assert ws.cell(2, 3).value == "a"
        wb.close()

    def test_with_bytesio(self):
        buf = io.BytesIO()
        df = pd.DataFrame({"X": [10, 20]})
        FastExcel(buf).sheet("Sheet1", df).save()
        buf.seek(0)
        assert buf.read(4) == XLSX_MAGIC


class TestDataFrameMultiSheet:
    def test_content_roundtrip(self, tmp_path):
        df1 = pd.DataFrame({"A": [1.0, 2.0]})
        df2 = pd.DataFrame({"B": [3.0, 4.0]})
        path = str(tmp_path / "df_multi.xlsx")
        FastExcel(path).sheet("Sheet1", df1).sheet("Sheet2", df2).save()

        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Sheet1", "Sheet2"]
        assert wb["Sheet1"].cell(2, 1).value == pytest.approx(1.0)
        assert wb["Sheet2"].cell(2, 1).value == pytest.approx(3.0)
        wb.close()


class TestDataFrameNumpyTypes:
    def test_numpy_int_written_as_number(self, tmp_path):
        """numpy.int64 should be written as Excel number, not string."""
        df = pd.DataFrame({"A": [1, 2, 3]})
        path = str(tmp_path / "np_int.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        val = ws.cell(2, 1).value
        assert isinstance(
            val, (int, float)
        ), f"Expected number, got {type(val).__name__}: {val!r}"
        assert val == pytest.approx(1.0)
        wb.close()

    def test_numpy_bool_written_as_boolean(self, tmp_path):
        """numpy.bool_ should be written as Excel boolean."""
        df = pd.DataFrame({"flag": [True, False]})
        path = str(tmp_path / "np_bool.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).value is True
        assert ws.cell(3, 1).value is False
        wb.close()
