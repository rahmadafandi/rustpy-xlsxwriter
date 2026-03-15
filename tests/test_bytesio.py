import io

import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel

from conftest import XLSX_MAGIC


class TestBytesIOSingleSheet:
    def test_roundtrip(self):
        buf = io.BytesIO()
        FastExcel(buf).sheet("Sheet1", [{"Name": "Alice", "Age": 30}]).save()

        buf.seek(0)
        assert buf.read(4) == XLSX_MAGIC

        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == 30
        wb.close()


class TestBytesIOMultiSheet:
    def test_roundtrip(self):
        buf = io.BytesIO()
        (
            FastExcel(buf)
            .sheet("Users", [{"Name": "Alice"}])
            .sheet("Items", [{"Item": "Laptop"}])
            .save()
        )

        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        assert wb.sheetnames == ["Users", "Items"]
        assert wb["Users"].cell(2, 1).value == "Alice"
        assert wb["Items"].cell(2, 1).value == "Laptop"
        wb.close()


class TestBytesIOErrors:
    def test_invalid_target_raises(self):
        with pytest.raises(TypeError):
            FastExcel(12345).sheet("Sheet1", [{"a": 1}]).save()
