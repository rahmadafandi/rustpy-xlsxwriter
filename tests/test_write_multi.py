import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel


class TestWriteMultiSheet:
    def test_two_sheets_content(self, tmp_path):
        path = str(tmp_path / "multi.xlsx")
        (
            FastExcel(path)
            .sheet("Users", [{"Name": "Alice"}, {"Name": "Bob"}])
            .sheet("Items", [{"SKU": "A1", "Price": 9.99}])
            .save()
        )

        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Users", "Items"]
        assert wb["Users"].cell(2, 1).value == "Alice"
        assert wb["Users"].cell(3, 1).value == "Bob"
        assert wb["Items"].cell(2, 1).value == "A1"
        assert wb["Items"].cell(2, 2).value == pytest.approx(9.99)
        wb.close()

    def test_many_sheets(self, tmp_path, small_records):
        path = str(tmp_path / "many_sheets.xlsx")
        writer = FastExcel(path)
        for i in range(10):
            writer.sheet(f"Sheet{i}", small_records)
        writer.save()

        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames) == 10
        wb.close()

    def test_invalid_sheet_name_raises(self):
        with pytest.raises(ValueError, match=r"Invalid sheet name"):
            (
                FastExcel("unused.xlsx")
                .sheet("Valid", [{"a": 1}])
                .sheet("Invalid*Name", [{"a": 1}])
                .save()
            )
