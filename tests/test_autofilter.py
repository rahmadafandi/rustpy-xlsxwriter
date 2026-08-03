"""Autofilter — filter dropdowns over the header row and its data.

The range depends on the final row count, which constant-memory mode only
knows once the last row is flushed, so it is applied after the data. These
tests pin the computed range for every data path and every row offset.
"""

import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel, write_worksheet, write_worksheets


def _records(n=5):
    return [{"name": f"n{i}", "score": i * 1.5, "count": i} for i in range(n)]


def _ref(path, sheet=None):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    return ws.auto_filter.ref


def test_off_by_default(tmp_path):
    path = tmp_path / "off.xlsx"
    write_worksheet(_records(), str(path))
    assert _ref(path) is None


def test_covers_header_and_all_data_rows(tmp_path):
    path = tmp_path / "on.xlsx"
    write_worksheet(_records(5), str(path), autofilter=True)
    # 3 columns, header on row 1 plus 5 data rows.
    assert _ref(path) == "A1:C6"


def test_range_follows_header_row(tmp_path):
    path = tmp_path / "offset.xlsx"
    write_worksheet(
        _records(4),
        str(path),
        header_row=2,
        merge_ranges=[(0, 0, 0, 2, "Banner")],
        autofilter=True,
    )
    # Header on sheet row 3, 4 data rows below it.
    assert _ref(path) == "A3:C7"


def test_empty_records_filters_header_only(tmp_path):
    path = tmp_path / "empty.xlsx"
    write_worksheet([], str(path), autofilter=True)
    # No headers were written, so there is nothing to filter.
    assert _ref(path) is None


def test_single_row(tmp_path):
    path = tmp_path / "one.xlsx"
    write_worksheet(_records(1), str(path), autofilter=True)
    assert _ref(path) == "A1:C2"


def test_combines_with_freeze_and_banding(tmp_path):
    path = tmp_path / "combo.xlsx"
    write_worksheet(
        _records(4),
        str(path),
        freeze_row=1,
        banded_rows="#F2F2F2",
        autofilter=True,
    )
    ws = openpyxl.load_workbook(path).active
    assert ws.auto_filter.ref == "A1:C5"
    assert ws.freeze_panes == "A2"
    assert ws.cell(row=3, column=1).fill.fgColor.rgb == "FFF2F2F2"


@pytest.mark.parametrize("frame", ["pandas", "polars"])
def test_dataframe_paths(tmp_path, frame):
    mod = pytest.importorskip(frame)
    df = mod.DataFrame({"a": [1, 2, 3], "b": ["x", "y", "z"]})
    path = tmp_path / f"{frame}.xlsx"
    write_worksheet(df, str(path), autofilter=True)
    assert _ref(path) == "A1:B4"


def test_dataframe_fallback_path(tmp_path):
    """The non-Arrow writer reports its row count separately."""
    from tests.test_row_layout import _FakeFrame

    df = _FakeFrame({"a": [1, 2, 3, 4]}, kinds=["i"])
    path = tmp_path / "fallback.xlsx"
    write_worksheet(df, str(path), autofilter=True)
    assert _ref(path) == "A1:A5"


def test_multi_sheet_is_per_sheet(tmp_path):
    path = tmp_path / "multi.xlsx"
    write_worksheets(
        [("Filtered", _records(3)), ("Plain", _records(3))],
        str(path),
        autofilter={"Filtered": True},
    )
    assert _ref(path, "Filtered") == "A1:C4"
    assert _ref(path, "Plain") is None


def test_multi_sheet_general_key(tmp_path):
    path = tmp_path / "general.xlsx"
    write_worksheets(
        [("A", _records(2)), ("B", _records(3))],
        str(path),
        autofilter={"general": True},
    )
    assert _ref(path, "A") == "A1:C3"
    assert _ref(path, "B") == "A1:C4"


def test_fastexcel_builder(tmp_path):
    path = tmp_path / "builder.xlsx"
    FastExcel(str(path)).sheet("S", _records(6), autofilter=True).save()
    assert _ref(path, "S") == "A1:C7"


def test_csv_warns_that_autofilter_is_dropped(tmp_path):
    path = tmp_path / "o.csv"
    with pytest.warns(UserWarning, match="autofilter"):
        FastExcel(str(path)).sheet("S", _records(2), autofilter=True).save()
