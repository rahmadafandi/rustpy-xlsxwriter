import io

import openpyxl

from rustpy_xlsxwriter import write_worksheet, write_worksheets

from conftest import XLSX_MAGIC


class TestWriteWorksheet:
    def test_write_to_file(self, tmp_path):
        path = str(tmp_path / "functional.xlsx")
        records = [{"X": 1, "Y": 2}, {"X": 3, "Y": 4}]
        write_worksheet(records, path, sheet_name="Data")

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.title == "Data"
        assert ws.cell(2, 1).value == 1
        assert ws.cell(3, 2).value == 4
        wb.close()

    def test_write_to_bytesio(self):
        buf = io.BytesIO()
        write_worksheet([{"A": "hello"}], buf)
        buf.seek(0)
        assert buf.read(4) == XLSX_MAGIC


class TestWriteWorksheets:
    def test_write_multiple_sheets(self, tmp_path):
        path = str(tmp_path / "multi_func.xlsx")
        write_worksheets(
            [{"Sheet1": [{"A": 1}]}, {"Sheet2": [{"B": 2}]}],
            path,
        )

        wb = openpyxl.load_workbook(path)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames
        assert wb["Sheet1"].cell(2, 1).value == 1
        assert wb["Sheet2"].cell(2, 1).value == 2
        wb.close()
