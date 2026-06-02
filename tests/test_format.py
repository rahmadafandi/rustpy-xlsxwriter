import openpyxl
import pytest

from rustpy_xlsxwriter import Format, write_worksheet

RECORDS = [{"name": "Alice", "score": 90}, {"name": "Bob", "score": 80}]


def load(path):
    return openpyxl.load_workbook(str(path)).active


class TestColumnFormatsRecords:
    def test_bold_by_name(self, tmp_path):
        p = tmp_path / "bold.xlsx"
        write_worksheet(
            RECORDS, p, column_formats={"name": Format().set_bold()}, autofit=False
        )
        ws = load(p)
        assert ws["A2"].font.bold is True
        assert ws["B2"].font.bold in (False, None)

    def test_bg_color_by_list(self, tmp_path):
        p = tmp_path / "bg.xlsx"
        write_worksheet(
            RECORDS, p,
            column_formats=[Format().set_background_color("#FFFF00"), None],
            autofit=False,
        )
        ws = load(p)
        assert ws["A2"].fill.fgColor.rgb.endswith("FFFF00")

    def test_named_color(self, tmp_path):
        p = tmp_path / "named.xlsx"
        write_worksheet(
            RECORDS, p, column_formats={"name": Format().set_font_color("red")},
            autofit=False,
        )
        ws = load(p)
        assert ws["A2"].font.color.rgb.endswith("FF0000")

    def test_chaining(self, tmp_path):
        # Chaining must return a Format usable downstream.
        p = tmp_path / "chain.xlsx"
        f = Format().set_bold().set_font_color("red").set_align("center")
        write_worksheet(RECORDS, p, column_formats={"name": f}, autofit=False)
        ws = load(p)
        assert ws["A2"].font.bold is True


class TestHeaderFormat:
    def test_header_format_applied(self, tmp_path):
        p = tmp_path / "hdr.xlsx"
        write_worksheet(
            RECORDS, p, header_format=Format().set_bold().set_font_color("#0000FF"),
            autofit=False,
        )
        ws = load(p)
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.color.rgb.endswith("0000FF")
        assert ws["B1"].font.bold is True

    def test_header_format_beats_bold_headers(self, tmp_path):
        p = tmp_path / "hdr2.xlsx"
        write_worksheet(
            RECORDS, p, bold_headers=True,
            header_format=Format().set_italic(),
            autofit=False,
        )
        ws = load(p)
        assert ws["A1"].font.italic is True


class TestNumericPrecedence:
    def test_column_num_format_beats_float_format(self, tmp_path):
        p = tmp_path / "num.xlsx"
        rows = [{"a": 1.5}, {"a": 2.5}]
        write_worksheet(
            rows, p, float_format="0.000",
            column_formats={"a": Format().set_num_format("0.00%")},
            autofit=False,
        )
        ws = load(p)
        assert ws["A2"].number_format == "0.00%"

    def test_no_override_keeps_float_format(self, tmp_path):
        p = tmp_path / "num2.xlsx"
        rows = [{"a": 1.5}]
        write_worksheet(rows, p, float_format="0.000", autofit=False)
        ws = load(p)
        assert ws["A2"].number_format == "0.000"


class TestDataFrameFormats:
    def test_pandas_column_format(self, tmp_path):
        pd = pytest.importorskip("pandas")
        p = tmp_path / "pd.xlsx"
        df = pd.DataFrame({"x": [1.0, 2.0], "y": ["a", "b"]})
        write_worksheet(
            df, p,
            column_formats={"x": Format().set_num_format("0.00%"),
                            "y": Format().set_bold()},
            header_format=Format().set_bold(),
            autofit=False,
        )
        ws = load(p)
        assert ws["A2"].number_format == "0.00%"
        assert ws["B2"].font.bold is True
        assert ws["A1"].font.bold is True

    def test_polars_column_format(self, tmp_path):
        pl = pytest.importorskip("polars")
        p = tmp_path / "pl.xlsx"
        df = pl.DataFrame({"x": [1.0, 2.0], "y": ["a", "b"]})
        write_worksheet(
            df, p, column_formats=[Format().set_num_format("0.00%"), Format().set_bold()],
            autofit=False,
        )
        ws = load(p)
        assert ws["A2"].number_format == "0.00%"
        assert ws["B2"].value == "a"
        assert ws["B2"].font.bold is True


class TestAllSetters:
    def test_font_attrs(self, tmp_path):
        p = tmp_path / "font.xlsx"
        f = (Format().set_font_name("Arial").set_font_size(14)
             .set_italic().set_font_strikethrough().set_underline("single"))
        write_worksheet(RECORDS, p, column_formats={"name": f}, autofit=False)
        c = load(p)["A2"]
        assert c.font.name == "Arial"
        assert c.font.sz == 14
        assert c.font.italic is True
        assert c.font.strike is True
        assert c.font.underline == "single"

    def test_border_sides(self, tmp_path):
        p = tmp_path / "border.xlsx"
        f = (Format().set_border("thin").set_border_top("double")
             .set_border_color("#FF0000"))
        write_worksheet(RECORDS, p, column_formats={"name": f}, autofit=False)
        c = load(p)["A2"]
        assert c.border.top.style == "double"
        assert c.border.left.style == "thin"

    def test_alignment_wrap(self, tmp_path):
        p = tmp_path / "align.xlsx"
        f = Format().set_align("center").set_align("vcenter").set_text_wrap()
        write_worksheet(RECORDS, p, column_formats={"name": f}, autofit=False)
        c = load(p)["A2"]
        assert c.alignment.horizontal == "center"
        assert c.alignment.vertical == "center"
        assert c.alignment.wrap_text is True

    def test_pattern_solid(self, tmp_path):
        p = tmp_path / "pat.xlsx"
        f = Format().set_pattern("solid").set_foreground_color("#00FF00")
        write_worksheet(RECORDS, p, column_formats={"name": f}, autofit=False)
        c = load(p)["A2"]
        assert c.fill.patternType == "solid"

    def test_underline_default(self, tmp_path):
        p = tmp_path / "ul.xlsx"
        f = Format().set_underline()  # defaults to single
        write_worksheet(RECORDS, p, column_formats={"name": f}, autofit=False)
        assert load(p)["A2"].font.underline == "single"


class TestMultiSheetFormats:
    def _ws(self, path, sheet):
        return openpyxl.load_workbook(str(path))[sheet]

    def test_keyed(self, tmp_path):
        from rustpy_xlsxwriter import write_worksheets
        p = tmp_path / "multi.xlsx"
        write_worksheets(
            [("Raw", RECORDS), ("Meta", RECORDS)],
            p, autofit=False,
            column_formats={"Meta": {"name": Format().set_bold()}},
            header_format={"general": Format().set_italic()},
        )
        raw = self._ws(p, "Raw")
        meta = self._ws(p, "Meta")
        assert raw["A1"].font.italic is True     # general header_format applies to all
        assert meta["A1"].font.italic is True
        assert meta["A2"].font.bold is True      # Meta-specific column_formats
        assert raw["A2"].font.bold in (False, None)  # Raw has no column_formats


class TestBuilderFormats:
    def test_builder(self, tmp_path):
        from rustpy_xlsxwriter import FastExcel
        p = tmp_path / "b.xlsx"
        (
            FastExcel(p, autofit=False)
            .sheet("S", RECORDS,
                   column_formats={"name": Format().set_bold()},
                   header_format=Format().set_italic())
            .save()
        )
        ws = openpyxl.load_workbook(str(p)).active
        assert ws["A1"].font.italic is True
        assert ws["A2"].font.bold is True

    def test_builder_multi(self, tmp_path):
        from rustpy_xlsxwriter import FastExcel
        p = tmp_path / "bm.xlsx"
        (
            FastExcel(p, autofit=False)
            .sheet("Raw", RECORDS)
            .sheet("Meta", RECORDS, column_formats=[Format().set_bold(), None])
            .save()
        )
        meta = openpyxl.load_workbook(str(p))["Meta"]
        assert meta["A2"].font.bold is True
        raw = openpyxl.load_workbook(str(p))["Raw"]
        assert raw["A2"].font.bold in (False, None)


class TestFormatErrors:
    def test_invalid_color_raises(self):
        with pytest.raises(ValueError, match="invalid color"):
            Format().set_font_color("notacolor")

    def test_invalid_align_raises(self):
        with pytest.raises(ValueError, match="invalid align"):
            Format().set_align("sideways")

    def test_invalid_border_raises(self):
        with pytest.raises(ValueError, match="invalid border"):
            Format().set_border("zigzag")

    def test_invalid_pattern_raises(self):
        with pytest.raises(ValueError):
            Format().set_pattern("plaid")

    def test_invalid_underline_raises(self):
        with pytest.raises(ValueError):
            Format().set_underline("wiggly")

    def test_non_format_value_raises(self, tmp_path):
        p = tmp_path / "bad.xlsx"
        with pytest.raises(TypeError):
            write_worksheet(RECORDS, p, column_formats={"name": "bold"}, autofit=False)

    def test_bad_spec_type_raises(self, tmp_path):
        p = tmp_path / "bad2.xlsx"
        with pytest.raises(ValueError, match="dict .* or a list"):
            write_worksheet(RECORDS, p, column_formats=42, autofit=False)

    def test_unknown_column_warns(self, tmp_path):
        p = tmp_path / "warn.xlsx"
        with pytest.warns(UserWarning, match="unknown column 'nope'"):
            write_worksheet(
                RECORDS, p, column_formats={"nope": Format().set_bold()}, autofit=False
            )

    def test_reused_format_two_columns(self, tmp_path):
        p = tmp_path / "reuse.xlsx"
        f = Format().set_bold()
        write_worksheet(RECORDS, p, column_formats={"name": f, "score": f}, autofit=False)
        ws = load(p)
        assert ws["A2"].font.bold is True
        assert ws["B2"].font.bold is True
