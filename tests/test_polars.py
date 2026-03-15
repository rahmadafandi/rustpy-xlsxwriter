import io
from datetime import date, datetime

import openpyxl
import polars as pl
import pytest

from rustpy_xlsxwriter import FastExcel

from conftest import XLSX_MAGIC


class TestPolarsSingleSheet:
    def test_content_roundtrip(self, tmp_path):
        df = pl.DataFrame(
            {"Name": ["Alice", "Bob"], "Age": [30, 25], "Score": [95.5, 88.0]}
        )
        path = str(tmp_path / "polars.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert [ws.cell(1, c).value for c in range(1, 4)] == ["Name", "Age", "Score"]
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == pytest.approx(30.0)
        assert ws.cell(2, 3).value == pytest.approx(95.5)
        assert ws.cell(3, 1).value == "Bob"
        wb.close()

    def test_empty_dataframe(self, tmp_path):
        df = pl.DataFrame({"A": [], "B": []})
        path = str(tmp_path / "polars_empty.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).value == "A"
        assert ws.cell(1, 2).value == "B"
        assert ws.cell(2, 1).value is None
        wb.close()

    def test_with_bytesio(self):
        buf = io.BytesIO()
        df = pl.DataFrame({"X": [10, 20]})
        FastExcel(buf).sheet("Sheet1", df).save()
        buf.seek(0)
        assert buf.read(4) == XLSX_MAGIC

    def test_mixed_column_types(self, tmp_path):
        df = pl.DataFrame(
            {
                "int_col": [1, 2],
                "float_col": [1.5, 2.5],
                "str_col": ["a", "b"],
                "bool_col": [True, False],
            }
        )
        path = str(tmp_path / "polars_mixed.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).value == pytest.approx(1.0)
        assert ws.cell(2, 2).value == pytest.approx(1.5)
        assert ws.cell(2, 3).value == "a"
        assert ws.cell(2, 4).value is True
        assert ws.cell(3, 4).value is False
        wb.close()


class TestPolarsMultiSheet:
    def test_content_roundtrip(self, tmp_path):
        df1 = pl.DataFrame({"A": [1.0, 2.0]})
        df2 = pl.DataFrame({"B": [3.0, 4.0]})
        path = str(tmp_path / "polars_multi.xlsx")
        FastExcel(path).sheet("Sheet1", df1).sheet("Sheet2", df2).save()

        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Sheet1", "Sheet2"]
        assert wb["Sheet1"].cell(2, 1).value == pytest.approx(1.0)
        assert wb["Sheet2"].cell(2, 1).value == pytest.approx(3.0)
        wb.close()


class TestPolarsDatetime:
    def test_datetime_values(self, tmp_path):
        df = pl.DataFrame({"ts": [datetime(2024, 6, 15, 10, 30, 45)]})
        path = str(tmp_path / "polars_dt.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        cell_val = ws.cell(2, 1).value
        assert isinstance(cell_val, datetime)
        assert cell_val.year == 2024
        assert cell_val.month == 6
        assert cell_val.day == 15
        wb.close()

    def test_date_values(self, tmp_path):
        df = pl.DataFrame({"d": [date(2024, 3, 20), date(1999, 12, 31)]})
        path = str(tmp_path / "polars_date.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        cell_val = ws.cell(2, 1).value
        assert isinstance(cell_val, datetime)
        assert cell_val.year == 2024
        assert cell_val.month == 3
        assert cell_val.day == 20
        wb.close()

    def test_custom_datetime_format(self, tmp_path):
        df = pl.DataFrame({"ts": [datetime(2024, 1, 1)]})
        path = str(tmp_path / "polars_dt_fmt.xlsx")
        FastExcel(path).format(datetime_format="dd/mm/yyyy").sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        assert wb.active.cell(2, 1).number_format == "dd/mm/yyyy"
        wb.close()


class TestPolarsNullable:
    def test_null_values(self, tmp_path):
        df = pl.DataFrame({"A": [1, None, 3], "B": ["x", None, "z"]})
        path = str(tmp_path / "polars_null.xlsx")
        FastExcel(path).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).value == pytest.approx(1.0)
        assert ws.cell(3, 1).value is None or ws.cell(3, 1).value == ""
        assert ws.cell(4, 1).value == pytest.approx(3.0)
        assert ws.cell(2, 2).value == "x"
        assert ws.cell(3, 2).value is None or ws.cell(3, 2).value == ""
        wb.close()


class TestPolarsStyling:
    def test_bold_headers(self, tmp_path):
        df = pl.DataFrame({"X": [1], "Y": [2]})
        path = str(tmp_path / "polars_bold.xlsx")
        FastExcel(path).format(bold_headers=True).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(1, 2).font.bold is True
        wb.close()

    def test_float_format(self, tmp_path):
        df = pl.DataFrame({"Val": [123.456]})
        path = str(tmp_path / "polars_float_fmt.xlsx")
        FastExcel(path).format(float_format="0.00").sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).value == pytest.approx(123.456)
        assert ws.cell(2, 1).number_format == "0.00"
        wb.close()

    def test_index_columns(self, tmp_path):
        df = pl.DataFrame({"ID": [1, 2], "Name": ["A", "B"]})
        path = str(tmp_path / "polars_idx.xlsx")
        FastExcel(path).format(index_columns=["ID"]).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).font.bold is True
        assert ws.cell(2, 2).font.bold is not True
        wb.close()
