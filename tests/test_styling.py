import io

import openpyxl
import pandas as pd
import pytest

from rustpy_xlsxwriter import FastExcel

from conftest import XLSX_MAGIC


class TestFloatFormat:
    def test_number_format_applied(self, tmp_path):
        records = [{"Value": 123.45678}]
        path = str(tmp_path / "float_fmt.xlsx")
        FastExcel(path).format(float_format="0.00").sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).value == pytest.approx(123.45678)
        assert ws.cell(2, 1).number_format == "0.00"
        wb.close()

    def test_with_bytesio(self):
        buf = io.BytesIO()
        df = pd.DataFrame({"Value": [123.45678]})
        FastExcel(buf).format(float_format="0.00").sheet("Sheet1", df).save()
        buf.seek(0)
        assert buf.read(4) == XLSX_MAGIC


class TestIndexColumns:
    def test_bold_applied_to_index_column(self, tmp_path):
        records = [{"ID": 1, "Name": "Alice"}, {"ID": 2, "Name": "Bob"}]
        path = str(tmp_path / "index_cols.xlsx")
        FastExcel(path).format(index_columns=["ID"]).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).font.bold is True
        assert ws.cell(2, 2).font.bold is not True
        wb.close()


class TestDatetimeFormat:
    def test_custom_datetime_format(self, tmp_path):
        from datetime import datetime

        records = [{"ts": datetime(2024, 6, 15, 10, 30, 0)}]
        path = str(tmp_path / "dt_fmt.xlsx")
        FastExcel(path).format(datetime_format="dd/mm/yyyy").sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        cell = ws.cell(2, 1)
        assert isinstance(cell.value, datetime)
        assert cell.number_format == "dd/mm/yyyy"
        wb.close()

    def test_default_datetime_format(self, tmp_path):
        from datetime import datetime

        records = [{"ts": datetime(2024, 1, 1)}]
        path = str(tmp_path / "dt_default.xlsx")
        FastExcel(path).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).number_format == "yyyy-mm-ddThh:mm:ss"
        wb.close()

    def test_dataframe_custom_datetime_format(self, tmp_path):
        from datetime import datetime

        df = pd.DataFrame({"ts": [datetime(2024, 6, 15)]})
        path = str(tmp_path / "df_dt_fmt.xlsx")
        FastExcel(path).format(datetime_format="yyyy-mm-dd").sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).number_format == "yyyy-mm-dd"
        wb.close()


class TestBoldHeaders:
    def test_bold_headers_enabled(self, tmp_path):
        records = [{"Name": "Alice", "Age": 30}]
        path = str(tmp_path / "bold_hdr.xlsx")
        FastExcel(path).format(bold_headers=True).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(1, 2).font.bold is True
        # Data cells should NOT be bold
        assert ws.cell(2, 1).font.bold is not True
        wb.close()

    def test_bold_headers_disabled_by_default(self, tmp_path):
        records = [{"Name": "Alice"}]
        path = str(tmp_path / "no_bold_hdr.xlsx")
        FastExcel(path).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).font.bold is not True
        wb.close()

    def test_bold_headers_with_dataframe(self, tmp_path):
        df = pd.DataFrame({"X": [1], "Y": [2]})
        path = str(tmp_path / "df_bold_hdr.xlsx")
        FastExcel(path).format(bold_headers=True).sheet("Sheet1", df).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(1, 2).font.bold is True
        wb.close()


class TestCombinedStyling:
    def test_float_format_index_columns_and_freeze(self, tmp_path):
        records = [{"ID": 1, "Score": 99.123}]
        path = str(tmp_path / "combined.xlsx")
        (
            FastExcel(path)
            .format(float_format="0.00", index_columns=["ID"])
            .freeze(row=1)
            .sheet("Data", records)
            .save()
        )

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.freeze_panes == "A2"
        assert ws.cell(2, 1).font.bold is True
        assert ws.cell(2, 2).number_format == "0.00"
        wb.close()
