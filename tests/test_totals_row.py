"""Totals row — aggregate formulas written below the data.

Like autofilter this lands after the data: the row sits underneath it and the
formula ranges depend on how many rows there turned out to be.
"""

import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel, Format, write_worksheet, write_worksheets


def _records(n=4):
    return [{"name": f"n{i}", "amount": i * 1.5, "qty": i} for i in range(n)]


def _sheet(path, name=None):
    wb = openpyxl.load_workbook(path)
    return wb[name] if name else wb.active


def test_off_by_default(tmp_path):
    path = tmp_path / "off.xlsx"
    write_worksheet(_records(), str(path))
    assert _sheet(path)["A6"].value is None


def test_sum_formula_covers_the_data_range(tmp_path):
    path = tmp_path / "sum.xlsx"
    write_worksheet(_records(4), str(path), totals_row={"amount": "sum"})
    ws = _sheet(path)
    # Header row 1, data rows 2-5, totals on row 6.
    assert ws["B6"].value == "=SUM(B2:B5)"
    assert ws["A6"].value is None


def test_label_and_format(tmp_path):
    path = tmp_path / "labelled.xlsx"
    write_worksheet(
        _records(3),
        str(path),
        totals_row={"amount": "sum", "qty": "sum"},
        totals_label="Total",
        totals_format=Format().set_bold().set_border_top("thin"),
    )
    ws = _sheet(path)
    assert ws["A5"].value == "Total"
    assert ws["B5"].value == "=SUM(B2:B4)"
    assert ws["C5"].value == "=SUM(C2:C4)"
    for ref in ("A5", "B5", "C5"):
        assert ws[ref].font.bold
        assert ws[ref].border.top.style == "thin"


@pytest.mark.parametrize(
    "aggregate,expected",
    [
        ("sum", "SUM"),
        ("average", "AVERAGE"),
        ("avg", "AVERAGE"),
        ("mean", "AVERAGE"),
        ("count", "COUNT"),
        ("min", "MIN"),
        ("max", "MAX"),
        ("product", "PRODUCT"),
        ("stdev", "STDEV"),
        ("SUM", "SUM"),
    ],
)
def test_supported_aggregates(tmp_path, aggregate, expected):
    path = tmp_path / f"{aggregate}.xlsx"
    write_worksheet(_records(3), str(path), totals_row={"amount": aggregate})
    assert _sheet(path)["B5"].value == f"={expected}(B2:B4)"


def test_unknown_aggregate_raises(tmp_path):
    with pytest.raises(ValueError, match="unknown aggregate 'summ'"):
        write_worksheet(
            _records(), str(tmp_path / "bad.xlsx"), totals_row={"amount": "summ"}
        )


def test_unknown_column_warns_and_skips(tmp_path):
    path = tmp_path / "warn.xlsx"
    with pytest.warns(UserWarning, match="unknown column 'nope'"):
        write_worksheet(
            _records(3),
            str(path),
            totals_row={"nope": "sum", "amount": "sum"},
        )
    # The valid column is still totalled.
    assert _sheet(path)["B5"].value == "=SUM(B2:B4)"


def test_label_colliding_with_a_formula_raises(tmp_path):
    with pytest.raises(ValueError, match="totals_label would overwrite"):
        write_worksheet(
            _records(),
            str(tmp_path / "clash.xlsx"),
            totals_row={"name": "count"},
            totals_label="Total",
        )


def test_no_data_means_no_totals_row(tmp_path):
    """A formula over an empty range would be invalid."""
    path = tmp_path / "empty.xlsx"
    write_worksheet([], str(path), totals_row={"amount": "sum"}, totals_label="Total")
    ws = _sheet(path)
    assert ws["A1"].value is None
    assert ws["A2"].value is None


def test_single_row(tmp_path):
    path = tmp_path / "one.xlsx"
    write_worksheet(_records(1), str(path), totals_row={"amount": "sum"})
    assert _sheet(path)["B3"].value == "=SUM(B2:B2)"


def test_range_follows_header_row(tmp_path):
    path = tmp_path / "offset.xlsx"
    write_worksheet(
        _records(3),
        str(path),
        header_row=2,
        merge_ranges=[(0, 0, 0, 2, "Banner")],
        totals_row={"amount": "sum"},
    )
    # Header on row 3, data rows 4-6, totals on row 7.
    assert _sheet(path)["B7"].value == "=SUM(B4:B6)"


def test_autofilter_range_excludes_the_totals_row(tmp_path):
    """Otherwise sorting or filtering would drag the total into the data."""
    path = tmp_path / "filtered.xlsx"
    write_worksheet(
        _records(4),
        str(path),
        autofilter=True,
        totals_row={"amount": "sum"},
    )
    ws = _sheet(path)
    assert ws.auto_filter.ref == "A1:C5"
    assert ws["B6"].value == "=SUM(B2:B5)"


def test_uncomputed_result_reads_as_none_not_zero(tmp_path):
    """The cached result must be empty, never a plausible-looking wrong total.

    rust_xlsxwriter defaults the cached result to 0, so a reader trusting the
    cache would see ``0`` as the sum. An empty result reads back as "not
    computed" — and per the crate's docs it is also what makes LibreOffice
    recalculate instead of showing the stale value.
    """
    path = tmp_path / "cached.xlsx"
    write_worksheet(_records(3), str(path), totals_row={"amount": "sum"})

    cached = openpyxl.load_workbook(path, data_only=True).active
    assert cached["B5"].value is None
    # The formula itself is intact.
    assert openpyxl.load_workbook(path).active["B5"].value == "=SUM(B2:B4)"


def test_data_cells_keep_their_values(tmp_path):
    """set_formula_result_default must not touch ordinary cells."""
    path = tmp_path / "data.xlsx"
    write_worksheet(_records(3), str(path), totals_row={"amount": "sum"})
    cached = openpyxl.load_workbook(path, data_only=True).active
    assert cached["B2"].value == 0.0
    assert cached["B3"].value == 1.5
    assert cached["A2"].value == "n0"


@pytest.mark.parametrize("frame", ["pandas", "polars"])
def test_dataframe_paths(tmp_path, frame):
    mod = pytest.importorskip(frame)
    df = mod.DataFrame({"name": ["a", "b", "c"], "amount": [1.0, 2.0, 3.0]})
    path = tmp_path / f"{frame}.xlsx"
    write_worksheet(df, str(path), totals_row={"amount": "sum"})
    assert _sheet(path)["B5"].value == "=SUM(B2:B4)"


def test_dataframe_fallback_path(tmp_path):
    from tests.test_row_layout import _FakeFrame

    df = _FakeFrame({"amount": [1.0, 2.0]}, kinds=["f"])
    path = tmp_path / "fallback.xlsx"
    write_worksheet(df, str(path), totals_row={"amount": "sum"})
    assert _sheet(path)["A4"].value == "=SUM(A2:A3)"


def test_column_letter_past_z(tmp_path):
    """Column 27 is AB — the letter conversion must not be a single char."""
    row = {f"c{i}": i for i in range(28)}
    path = tmp_path / "wide.xlsx"
    write_worksheet([row, row], str(path), totals_row={"c27": "sum"})
    assert _sheet(path)["AB4"].value == "=SUM(AB2:AB3)"


def test_multi_sheet_is_per_sheet(tmp_path):
    path = tmp_path / "multi.xlsx"
    write_worksheets(
        [("Totalled", _records(3)), ("Plain", _records(3))],
        str(path),
        totals_row={"Totalled": {"amount": "sum"}},
        totals_label={"Totalled": "Total"},
    )
    assert _sheet(path, "Totalled")["B5"].value == "=SUM(B2:B4)"
    assert _sheet(path, "Totalled")["A5"].value == "Total"
    assert _sheet(path, "Plain")["A5"].value is None


def test_fastexcel_builder(tmp_path):
    path = tmp_path / "builder.xlsx"
    (
        FastExcel(str(path))
        .sheet(
            "S",
            _records(5),
            totals_row={"amount": "average"},
            totals_label="Mean",
            totals_format=Format().set_bold(),
        )
        .save()
    )
    ws = _sheet(path, "S")
    assert ws["A7"].value == "Mean"
    assert ws["B7"].value == "=AVERAGE(B2:B6)"


def test_csv_warns_that_totals_are_dropped(tmp_path):
    with pytest.warns(UserWarning, match="totals_row"):
        FastExcel(str(tmp_path / "o.csv")).sheet(
            "S", _records(3), totals_row={"amount": "sum"}
        ).save()
