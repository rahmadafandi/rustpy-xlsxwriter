"""Arbitrary formulas via ``formula_columns`` and free-form ``totals_row``.

Formula text is handed to ``rust_xlsxwriter`` verbatim, which rewrites 161
"future" and dynamic-array functions with an ``_xlfn.`` prefix and the right
XML shape. These tests check the text survives that trip intact — see
``test_formula_recalc.py`` for the values a real spreadsheet engine computes.
"""

import re
import zipfile

import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel, Format, write_worksheet, write_worksheets


def _rows(n=3):
    return [{"qty": i, "price": i * 2.0} for i in range(1, n + 1)]


def _sheet(path, name=None):
    wb = openpyxl.load_workbook(path)
    return wb[name] if name else wb.active


def _formula_xml(path, cell):
    """The raw <f> text, which is where the _xlfn. rewriting shows up."""
    xml = zipfile.ZipFile(path).read("xl/worksheets/sheet1.xml").decode()
    match = re.search(rf'<c r="{cell}"[^>]*>(?:<f[^>]*>(.*?)</f>)?', xml)
    return match.group(1) if match and match.group(1) else None


# --- formula columns ------------------------------------------------------


def test_appends_a_computed_column(tmp_path):
    path = tmp_path / "calc.xlsx"
    write_worksheet(_rows(3), str(path), formula_columns={"total": "=A{row}*B{row}"})

    ws = _sheet(path)
    assert [c.value for c in ws[1]] == ["qty", "price", "total"]
    assert ws["C2"].value == "=A2*B2"
    assert ws["C3"].value == "=A3*B3"
    assert ws["C4"].value == "=A4*B4"


def test_multiple_columns_keep_their_order(tmp_path):
    path = tmp_path / "multi.xlsx"
    write_worksheet(
        _rows(2),
        str(path),
        formula_columns={"total": "=A{row}*B{row}", "half": "=C{row}/2"},
    )
    ws = _sheet(path)
    assert [c.value for c in ws[1]] == ["qty", "price", "total", "half"]
    assert ws["C2"].value == "=A2*B2"
    assert ws["D2"].value == "=C2/2"


def test_first_placeholder(tmp_path):
    path = tmp_path / "first.xlsx"
    write_worksheet(
        _rows(3), str(path), formula_columns={"cum": "=SUM(B${first}:B{row})"}
    )
    ws = _sheet(path)
    assert ws["C2"].value == "=SUM(B$2:B2)"
    assert ws["C4"].value == "=SUM(B$2:B4)"


def test_last_placeholder_raises_with_a_useful_message(tmp_path):
    """It cannot be resolved while rows are still streaming."""
    with pytest.raises(ValueError, match=r"\{last\}.*still streaming"):
        write_worksheet(
            _rows(), str(tmp_path / "bad.xlsx"),
            formula_columns={"share": "=B{row}/SUM(B{first}:B{last})"},
        )


def test_empty_formula_raises(tmp_path):
    with pytest.raises(ValueError, match="is empty"):
        write_worksheet(
            _rows(), str(tmp_path / "bad.xlsx"), formula_columns={"x": "  "}
        )


def test_non_string_formula_raises(tmp_path):
    with pytest.raises(ValueError, match="must be a formula string"):
        write_worksheet(
            _rows(), str(tmp_path / "bad.xlsx"), formula_columns={"x": 42}
        )


def test_formula_columns_shift_with_header_row(tmp_path):
    path = tmp_path / "offset.xlsx"
    write_worksheet(
        _rows(2),
        str(path),
        header_row=1,
        merge_ranges=[(0, 0, 0, 1, "Banner")],
        formula_columns={"total": "=A{row}*B{row}"},
    )
    ws = _sheet(path)
    assert ws["C2"].value == "total"
    assert ws["C3"].value == "=A3*B3"


def test_autofilter_and_totals_cover_the_computed_column(tmp_path):
    path = tmp_path / "combo.xlsx"
    write_worksheet(
        _rows(3),
        str(path),
        formula_columns={"total": "=A{row}*B{row}"},
        autofilter=True,
        totals_row={"total": "sum"},
    )
    ws = _sheet(path)
    assert ws.auto_filter.ref == "A1:C4"
    assert ws["C5"].value == "=SUM(C2:C4)"


# --- the formula language itself -----------------------------------------


@pytest.mark.parametrize(
    "formula",
    [
        pytest.param("=IF(A{row}>1,SUM(B$2:B{row}),0)", id="nested-if-sum"),
        pytest.param('=IF(A{row}>1,IF(A{row}>2,"hi","mid"),"lo")', id="double-if"),
        pytest.param('=SUMIFS(B:B,A:A,">1")', id="sumifs"),
        pytest.param("=ROUND(B{row}/A{row},2)", id="round-div"),
        pytest.param("=SUMPRODUCT((A$2:A$4>1)*(B$2:B$4))", id="sumproduct"),
        pytest.param("=VLOOKUP(A{row},A:B,2,FALSE)", id="vlookup"),
        pytest.param('=A{row}&" x "&B{row}', id="concat-operator"),
        pytest.param("=INDEX(B:B,MATCH(A{row},A:A,0))", id="index-match"),
    ],
)
def test_classic_formulas_pass_through_verbatim(tmp_path, formula):
    path = tmp_path / "classic.xlsx"
    write_worksheet(_rows(3), str(path), formula_columns={"f": formula})
    assert _sheet(path)["C2"].value == formula.replace("{row}", "2")


@pytest.mark.parametrize(
    "formula,expected_prefix",
    [
        pytest.param('=TEXTJOIN(",",TRUE,A$2:A$4)', "_xlfn.TEXTJOIN", id="textjoin"),
        pytest.param('=IFS(A{row}>2,"hi",TRUE,"lo")', "_xlfn.IFS", id="ifs"),
        pytest.param("=CONCAT(A$2:A$4)", "_xlfn.CONCAT", id="concat"),
        pytest.param('=MAXIFS(B:B,A:A,">1")', "_xlfn.MAXIFS", id="maxifs"),
        pytest.param("=STDEV.P(B$2:B$4)", "_xlfn.STDEV.P", id="stdev-p"),
        pytest.param("=IFNA(A{row},0)", "_xlfn.IFNA", id="ifna"),
    ],
)
def test_future_functions_get_the_xlfn_prefix(tmp_path, formula, expected_prefix):
    """Excel needs these annotated in the file; the crate does it for us."""
    path = tmp_path / "future.xlsx"
    write_worksheet(_rows(3), str(path), formula_columns={"f": formula})
    assert expected_prefix in _formula_xml(path, "C2")


@pytest.mark.parametrize(
    "formula,expected_prefix",
    [
        pytest.param("=XLOOKUP(A{row},A:A,B:B)", "_xlfn.XLOOKUP", id="xlookup"),
        pytest.param("=UNIQUE(A$2:A$4)", "_xlfn.UNIQUE", id="unique"),
        pytest.param("=SORT(A$2:A$4)", "_xlfn._xlws.SORT", id="sort-needs-xlws"),
        pytest.param('=FILTER(A$2:A$4,B$2:B$4>2)', "_xlfn._xlws.FILTER", id="filter"),
    ],
)
def test_dynamic_array_functions(tmp_path, formula, expected_prefix):
    path = tmp_path / "dynamic.xlsx"
    write_worksheet(_rows(3), str(path), formula_columns={"f": formula})

    xml = zipfile.ZipFile(path).read("xl/worksheets/sheet1.xml").decode()
    cell = re.search(r'<c r="C2".*?</c>', xml).group(0)
    assert expected_prefix in cell
    # Dynamic arrays need the array shape plus the metadata part.
    assert 't="array"' in cell
    assert 'cm="1"' in cell
    assert "xl/metadata.xml" in zipfile.ZipFile(path).namelist()


def test_special_characters_are_xml_escaped(tmp_path):
    path = tmp_path / "escape.xlsx"
    write_worksheet(
        _rows(2),
        str(path),
        formula_columns={"f": '=IF(A{row}<2,"a&b",">c")'},
    )
    raw = _formula_xml(path, "C2")
    assert "&lt;" in raw and "&gt;" in raw and "&amp;" in raw
    # openpyxl unescapes it back to the original text.
    assert _sheet(path)["C2"].value == '=IF(A2<2,"a&b",">c")'


def test_unknown_function_is_written_unchanged(tmp_path):
    """Structure is validated; function names deliberately are not.

    LAMBDA and LET bind their own names, workbooks carry user-defined
    functions, and Excel keeps adding to the list — a name whitelist would
    reject valid formulas, which is worse than letting a typo through to a
    ``#NAME?`` in one cell.
    """
    path = tmp_path / "invalid.xlsx"
    write_worksheet(_rows(2), str(path), formula_columns={"f": "=NOTAFUNC(A{row})"})
    assert _sheet(path)["C2"].value == "=NOTAFUNC(A2)"


# --- free-form totals -----------------------------------------------------


def test_totals_row_accepts_a_raw_formula(tmp_path):
    path = tmp_path / "totals.xlsx"
    write_worksheet(
        _rows(3),
        str(path),
        totals_row={"price": "=MAX({col}{first}:{col}{last})"},
    )
    assert _sheet(path)["B5"].value == "=MAX(B2:B4)"


def test_totals_formula_and_aggregate_mix(tmp_path):
    path = tmp_path / "mixed.xlsx"
    write_worksheet(
        _rows(3),
        str(path),
        totals_row={
            "qty": "sum",
            "price": "=ROUND(AVERAGE({col}{first}:{col}{last}),2)",
        },
    )
    ws = _sheet(path)
    assert ws["A5"].value == "=SUM(A2:A4)"
    assert ws["B5"].value == "=ROUND(AVERAGE(B2:B4),2)"


def test_a_typo_still_raises_rather_than_becoming_a_label(tmp_path):
    """Only a leading '=' opts into raw-formula mode."""
    with pytest.raises(ValueError, match="unknown aggregate 'summ'"):
        write_worksheet(
            _rows(), str(tmp_path / "bad.xlsx"), totals_row={"qty": "summ"}
        )


# --- all four data paths --------------------------------------------------


@pytest.mark.parametrize("frame", ["pandas", "polars"])
def test_dataframe_paths(tmp_path, frame):
    mod = pytest.importorskip(frame)
    df = mod.DataFrame({"qty": [1, 2, 3], "price": [2.0, 4.0, 6.0]})
    path = tmp_path / f"{frame}.xlsx"
    write_worksheet(df, str(path), formula_columns={"total": "=A{row}*B{row}"})

    ws = _sheet(path)
    assert [c.value for c in ws[1]] == ["qty", "price", "total"]
    assert ws["C2"].value == "=A2*B2"
    assert ws["C4"].value == "=A4*B4"


def test_dataframe_fallback_path(tmp_path):
    from tests.test_row_layout import _FakeFrame

    df = _FakeFrame({"qty": [1, 2], "price": [2.0, 4.0]}, kinds=["i", "f"])
    path = tmp_path / "fallback.xlsx"
    write_worksheet(df, str(path), formula_columns={"total": "=A{row}*B{row}"})

    ws = _sheet(path)
    assert [c.value for c in ws[1]] == ["qty", "price", "total"]
    assert ws["C3"].value == "=A3*B3"


def test_generator_input(tmp_path):
    """Streaming input is the reason {last} is unavailable — check it works."""
    path = tmp_path / "gen.xlsx"
    write_worksheet(
        (r for r in _rows(4)), str(path), formula_columns={"total": "=A{row}*B{row}"}
    )
    ws = _sheet(path)
    assert ws["C5"].value == "=A5*B5"


def test_multi_sheet_is_per_sheet(tmp_path):
    path = tmp_path / "multi.xlsx"
    write_worksheets(
        [("Calc", _rows(2)), ("Plain", _rows(2))],
        str(path),
        formula_columns={"Calc": {"total": "=A{row}*B{row}"}},
    )
    wb = openpyxl.load_workbook(path)
    assert wb["Calc"]["C2"].value == "=A2*B2"
    assert wb["Plain"]["C2"].value is None


def test_fastexcel_builder(tmp_path):
    path = tmp_path / "builder.xlsx"
    (
        FastExcel(str(path))
        .sheet(
            "S",
            _rows(3),
            formula_columns={"total": "=A{row}*B{row}"},
            header_format=Format().set_bold(),
        )
        .save()
    )
    ws = _sheet(path, "S")
    assert ws["C1"].value == "total"
    assert ws["C1"].font.bold
    assert ws["C2"].value == "=A2*B2"


def test_csv_warns_that_formula_columns_are_dropped(tmp_path):
    with pytest.warns(UserWarning, match="formula_columns"):
        FastExcel(str(tmp_path / "o.csv")).sheet(
            "S", _rows(2), formula_columns={"total": "=A{row}*B{row}"}
        ).save()


# --- structural validation ------------------------------------------------
# Malformed formulas do not corrupt the file — every case below opens fine in
# LibreOffice and shows an error value in the cell. Validation only moves the
# discovery from "when someone opens the report" to "when the export runs", so
# it is deliberately limited to structure. Rejecting a formula Excel would have
# accepted is worse than passing a broken one through.


@pytest.mark.parametrize(
    "formula,problem",
    [
        pytest.param("=SUM(A2:A5", "unclosed", id="unclosed-paren"),
        pytest.param("=SUM((A2:A5)", "unclosed", id="one-of-two-unclosed"),
        pytest.param("=SUM(A2:A5))", "never opened", id="extra-close"),
        pytest.param('=IF(A2>1,"yes,"no")', "double quote", id="unbalanced-quote"),
        pytest.param("='Sheet1!A1", "single quote", id="unbalanced-sheet-quote"),
        pytest.param("=", "empty", id="bare-equals"),
        pytest.param("=   ", "empty", id="whitespace-only"),
    ],
)
def test_malformed_formulas_are_rejected(tmp_path, formula, problem):
    with pytest.raises(ValueError, match=problem):
        write_worksheet(
            _rows(), str(tmp_path / "bad.xlsx"), formula_columns={"f": formula}
        )


@pytest.mark.parametrize(
    "formula",
    [
        pytest.param('=IF(A2>1,"a)b","c(d")', id="parens-inside-string"),
        pytest.param("='Sheet (1)'!A1", id="parens-inside-sheet-name"),
        pytest.param('=A2&""""&B2', id="escaped-quote"),
        pytest.param('=CONCAT("it\'s",A2)', id="apostrophe-inside-string"),
        pytest.param("=NOTAFUNC(A2)", id="unknown-function-still-allowed"),
        pytest.param("SUM(A2:A3)", id="no-leading-equals"),
        pytest.param("=SUMPRODUCT((A2:A5>1)*(B2:B5))", id="nested-parens"),
    ],
)
def test_valid_formulas_are_not_rejected(tmp_path, formula):
    """False positives are the failure mode that matters here."""
    write_worksheet(
        _rows(), str(tmp_path / "ok.xlsx"), formula_columns={"f": formula}
    )


def test_totals_formula_is_validated_too(tmp_path):
    with pytest.raises(ValueError, match="totals_row\\['qty'\\].*unclosed"):
        write_worksheet(
            _rows(),
            str(tmp_path / "bad.xlsx"),
            totals_row={"qty": "=SUM({col}{first}:{col}{last}"},
        )


def test_error_message_names_the_column_and_shows_the_formula(tmp_path):
    with pytest.raises(ValueError) as excinfo:
        write_worksheet(
            _rows(), str(tmp_path / "bad.xlsx"), formula_columns={"margin": "=A2/("}
        )
    message = str(excinfo.value)
    assert "formula_columns['margin']" in message
    assert "=A2/(" in message
