import os
from datetime import date, datetime

import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel


class TestWriteSingleSheet:
    def test_basic_write_and_read_back(self, tmp_path):
        records = [
            {"Name": "Alice", "Age": 30, "Score": 95.5},
            {"Name": "Bob", "Age": 25, "Score": 88.0},
        ]
        path = str(tmp_path / "basic.xlsx")
        FastExcel(path).sheet("Data", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.title == "Data"
        assert [ws.cell(1, c).value for c in range(1, 4)] == ["Name", "Age", "Score"]
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == 30
        assert ws.cell(2, 3).value == 95.5
        assert ws.cell(3, 1).value == "Bob"
        assert ws.cell(3, 2).value == 25
        assert ws.cell(3, 3).value == 88.0
        wb.close()

    def test_none_values(self, tmp_path):
        records = [{"A": "hello", "B": None}, {"A": None, "B": "world"}]
        path = str(tmp_path / "nones.xlsx")
        FastExcel(path).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 2).value is None or ws.cell(2, 2).value == ""
        wb.close()

    def test_boolean_values(self, tmp_path):
        """Booleans should be written as Excel booleans, not integers."""
        records = [{"flag": True}, {"flag": False}]
        path = str(tmp_path / "bools.xlsx")
        FastExcel(path).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).value is True
        assert ws.cell(3, 1).value is False
        assert isinstance(ws.cell(2, 1).value, bool)
        assert isinstance(ws.cell(3, 1).value, bool)
        wb.close()

    def test_mixed_types(self, tmp_path):
        records = [
            {
                "str": "hello",
                "int": 42,
                "float": 3.14,
                "bool": True,
                "none": None,
            }
        ]
        path = str(tmp_path / "mixed.xlsx")
        FastExcel(path).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(2, 1).value == "hello"
        assert ws.cell(2, 2).value == 42
        assert ws.cell(2, 3).value == pytest.approx(3.14)
        assert ws.cell(2, 4).value is True
        wb.close()

    def test_datetime_values(self, tmp_path):
        dt = datetime(2024, 6, 15, 10, 30, 45)
        records = [{"ts": dt}]
        path = str(tmp_path / "datetime.xlsx")
        FastExcel(path).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        cell_val = ws.cell(2, 1).value
        assert isinstance(cell_val, datetime)
        assert cell_val.year == 2024
        assert cell_val.month == 6
        assert cell_val.day == 15
        assert cell_val.hour == 10
        assert cell_val.minute == 30
        assert cell_val.second == 45
        wb.close()

    def test_date_values(self, tmp_path):
        """date objects should be written as Excel dates."""
        d = date(2024, 3, 20)
        records = [{"d": d}]
        path = str(tmp_path / "date.xlsx")
        FastExcel(path).sheet("Sheet1", records).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        cell_val = ws.cell(2, 1).value
        assert isinstance(cell_val, datetime)
        assert cell_val.year == 2024
        assert cell_val.month == 3
        assert cell_val.day == 20
        wb.close()

    def test_empty_records(self, tmp_path):
        path = str(tmp_path / "empty.xlsx")
        FastExcel(path).sheet("Sheet1", []).save()
        assert os.path.exists(path)

    def test_generator_input(self, tmp_path):
        """Generators should work for memory-efficient streaming."""

        def gen():
            for i in range(100):
                yield {"id": i, "value": f"row_{i}"}

        path = str(tmp_path / "generator.xlsx")
        FastExcel(path).sheet("Data", gen()).save()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).value == "id"
        assert ws.cell(1, 2).value == "value"
        assert ws.cell(2, 1).value == 0
        assert ws.cell(2, 2).value == "row_0"
        assert ws.cell(101, 1).value == 99
        wb.close()

    def test_invalid_sheet_name_raises(self):
        with pytest.raises(ValueError, match=r"Invalid sheet name"):
            FastExcel("unused.xlsx").sheet("Test[]", [{"a": 1}]).save()

    def test_no_sheets_raises(self):
        with pytest.raises(ValueError, match=r"No sheets added"):
            FastExcel("unused.xlsx").save()

    def test_context_manager(self, tmp_path):
        """FastExcel should support 'with' statement for auto-save."""
        path = str(tmp_path / "ctx.xlsx")
        with FastExcel(path) as f:
            f.sheet("Sheet1", [{"Name": "Alice"}])

        wb = openpyxl.load_workbook(path)
        assert wb.active.cell(2, 1).value == "Alice"
        wb.close()

    def test_context_manager_no_save_on_exception(self, tmp_path):
        """Context manager should NOT save if an exception occurred."""
        path = str(tmp_path / "ctx_err.xlsx")
        with pytest.raises(RuntimeError):
            with FastExcel(path) as f:
                f.sheet("Sheet1", [{"A": 1}])
                raise RuntimeError("abort")

        assert not os.path.exists(path)

    def test_autofit_disabled(self, tmp_path):
        """autofit=False should still produce a valid file."""
        path = str(tmp_path / "no_autofit.xlsx")
        FastExcel(path, autofit=False).sheet("Sheet1", [{"Col": "data"}]).save()

        wb = openpyxl.load_workbook(path)
        assert wb.active.cell(2, 1).value == "data"
        wb.close()
