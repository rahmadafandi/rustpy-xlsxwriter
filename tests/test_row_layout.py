"""Merged cells, row heights, row formats and alternating row colours."""

import datetime

import openpyxl
import pytest

from rustpy_xlsxwriter import FastExcel, Format, write_worksheet, write_worksheets

BAND = "#F2F2F2"
BAND_ARGB = "FFF2F2F2"
NO_FILL = "00000000"


def _records(n=6):
    return [
        {
            "name": f"n{i}",
            "score": i + 0.5,
            "count": i,
            "ok": i % 2 == 0,
            "when": datetime.datetime(2026, 1, i + 1, 12, 0, 0),
        }
        for i in range(n)
    ]


def _fills(ws, row, ncols):
    return [ws.cell(row=row, column=c).fill.fgColor.rgb for c in range(1, ncols + 1)]


# --- merged cells ---------------------------------------------------------


def test_merge_range_banner_above_header(tmp_path):
    path = tmp_path / "crosstab.xlsx"
    banner = Format().set_bold().set_align("center")
    write_worksheet(
        [{"region": "North", "male": 1, "female": 2}],
        str(path),
        sheet_name="X",
        header_row=1,
        merge_ranges=[(0, 1, 0, 2, "Gender", banner)],
    )

    ws = openpyxl.load_workbook(path)["X"]
    assert [str(r) for r in ws.merged_cells.ranges] == ["B1:C1"]
    assert ws["B1"].value == "Gender"
    # Header pushed down to row 2 (1-based), data to row 3.
    assert [c.value for c in ws[2]] == ["region", "male", "female"]
    assert [c.value for c in ws[3]] == ["North", 1, 2]


def test_merge_without_format_is_allowed(tmp_path):
    path = tmp_path / "plain-merge.xlsx"
    write_worksheet(
        [{"a": 1}],
        str(path),
        header_row=1,
        merge_ranges=[(0, 0, 0, 1, "Banner")],
    )
    ws = openpyxl.load_workbook(path).active
    assert ws["A1"].value == "Banner"


def test_merge_overlapping_header_raises(tmp_path):
    """The silent-data-loss case: constant memory cannot merge a flushed row."""
    with pytest.raises(ValueError, match="header row is 0"):
        write_worksheet(
            [{"a": 1}],
            str(tmp_path / "bad.xlsx"),
            merge_ranges=[(0, 0, 0, 1, "Banner")],
        )


def test_merge_reaching_data_rows_raises(tmp_path):
    with pytest.raises(ValueError, match="raise header_row to at least 3"):
        write_worksheet(
            [{"a": 1}],
            str(tmp_path / "bad.xlsx"),
            header_row=1,
            merge_ranges=[(0, 0, 2, 1, "Tall banner")],
        )


def test_inverted_merge_raises(tmp_path):
    with pytest.raises(ValueError, match="inverted"):
        write_worksheet(
            [{"a": 1}],
            str(tmp_path / "bad.xlsx"),
            header_row=2,
            merge_ranges=[(1, 0, 0, 1, "Backwards")],
        )


def test_merge_wrong_arity_raises(tmp_path):
    with pytest.raises(ValueError, match="5 or 6 items"):
        write_worksheet(
            [{"a": 1}],
            str(tmp_path / "bad.xlsx"),
            header_row=1,
            merge_ranges=[(0, 0, 0, 1)],
        )


# --- row heights and row formats -----------------------------------------


def test_row_heights_and_row_formats(tmp_path):
    path = tmp_path / "rows.xlsx"
    write_worksheet(
        _records(3),
        str(path),
        row_heights={0: 40.0, 2: 30.0},
        row_formats={0: Format().set_border_bottom("thin")},
    )

    ws = openpyxl.load_workbook(path).active
    assert ws.row_dimensions[1].height == pytest.approx(40.0, abs=0.5)
    assert ws.row_dimensions[3].height == pytest.approx(30.0, abs=0.5)
    assert ws["A1"].border.bottom.style == "thin"


def test_row_height_survives_row_format_on_same_row(tmp_path):
    """`set_row_format` resets a row's height unless the height lands first."""
    path = tmp_path / "both.xlsx"
    write_worksheet(
        _records(2),
        str(path),
        row_heights={0: 44.0},
        row_formats={0: Format().set_bold()},
    )
    ws = openpyxl.load_workbook(path).active
    assert ws.row_dimensions[1].height == pytest.approx(44.0, abs=0.5)
    assert ws["A1"].font.bold


def test_negative_row_height_raises(tmp_path):
    with pytest.raises(ValueError, match="must not be negative"):
        write_worksheet([{"a": 1}], str(tmp_path / "bad.xlsx"), row_heights={0: -5})


def test_row_formats_rejects_non_format(tmp_path):
    with pytest.raises(ValueError, match="must be Format objects"):
        write_worksheet([{"a": 1}], str(tmp_path / "bad.xlsx"), row_formats={0: "bold"})


# --- banded rows ----------------------------------------------------------


def test_banded_rows_shade_every_other_data_row(tmp_path):
    path = tmp_path / "banded.xlsx"
    write_worksheet(_records(4), str(path), banded_rows=BAND)

    ws = openpyxl.load_workbook(path).active
    assert _fills(ws, 1, 5) == [NO_FILL] * 5  # header untouched
    assert _fills(ws, 2, 5) == [NO_FILL] * 5  # first data row unshaded
    assert _fills(ws, 3, 5) == [BAND_ARGB] * 5
    assert _fills(ws, 4, 5) == [NO_FILL] * 5
    assert _fills(ws, 5, 5) == [BAND_ARGB] * 5


def test_banding_covers_formatted_and_typed_cells(tmp_path):
    """The regression this feature exists for.

    A cell written with its own format ignores the row's, so banding done via
    ``set_row_format`` leaves holes in exactly the columns that carry a number
    format. Every column here has a different write path — float, int, bool,
    datetime, string, plus an explicit column format.
    """
    path = tmp_path / "holes.xlsx"
    write_worksheet(
        _records(4),
        str(path),
        float_format="0.00",
        banded_rows=BAND,
        column_formats={"count": Format().set_num_format("#,##0")},
    )

    ws = openpyxl.load_workbook(path).active
    for row in (3, 5):
        assert _fills(ws, row, 5) == [BAND_ARGB] * 5, f"hole in row {row}"


def test_banding_preserves_number_formats(tmp_path):
    path = tmp_path / "fmt.xlsx"
    write_worksheet(
        _records(4),
        str(path),
        float_format="0.00",
        banded_rows=BAND,
        column_formats={"count": Format().set_num_format("#,##0")},
    )
    ws = openpyxl.load_workbook(path).active
    # Column formats survive on both plain and banded rows.
    for row in (2, 3):
        assert ws.cell(row=row, column=2).number_format == "0.00"
        assert ws.cell(row=row, column=3).number_format == "#,##0"


def test_banding_preserves_values(tmp_path):
    path = tmp_path / "vals.xlsx"
    records = _records(4)
    write_worksheet(records, str(path), banded_rows=BAND)

    ws = openpyxl.load_workbook(path).active
    for idx, rec in enumerate(records):
        row = idx + 2
        assert ws.cell(row=row, column=1).value == rec["name"]
        assert ws.cell(row=row, column=2).value == pytest.approx(rec["score"])
        assert ws.cell(row=row, column=3).value == rec["count"]
        assert ws.cell(row=row, column=4).value == rec["ok"]
        assert ws.cell(row=row, column=5).value == rec["when"]


def test_banding_offset_by_header_row(tmp_path):
    """Banding counts from the first data row, not from sheet row 0."""
    path = tmp_path / "offset.xlsx"
    write_worksheet(
        _records(4),
        str(path),
        header_row=2,
        merge_ranges=[(0, 0, 0, 1, "Banner")],
        banded_rows=BAND,
    )
    ws = openpyxl.load_workbook(path).active
    assert _fills(ws, 4, 5) == [NO_FILL] * 5  # first data row
    assert _fills(ws, 5, 5) == [BAND_ARGB] * 5


def test_bad_band_colour_raises(tmp_path):
    with pytest.raises(ValueError):
        write_worksheet([{"a": 1}], str(tmp_path / "bad.xlsx"), banded_rows="not-a-colour")


# --- builder + multi-sheet ------------------------------------------------


def test_fastexcel_row_layout(tmp_path):
    path = tmp_path / "builder.xlsx"
    (
        FastExcel(str(path))
        .sheet(
            "Report",
            _records(4),
            header_row=1,
            merge_ranges=[(0, 1, 0, 2, "Metrics")],
            row_heights={1: 28.0},
            row_formats={1: Format().set_border_bottom("thin")},
            banded_rows=BAND,
        )
        .save()
    )

    ws = openpyxl.load_workbook(path)["Report"]
    assert [str(r) for r in ws.merged_cells.ranges] == ["B1:C1"]
    assert ws.row_dimensions[2].height == pytest.approx(28.0, abs=0.5)
    assert ws["A2"].border.bottom.style == "thin"
    assert _fills(ws, 4, 5) == [BAND_ARGB] * 5


def test_multi_sheet_layout_is_per_sheet(tmp_path):
    path = tmp_path / "multi.xlsx"
    write_worksheets(
        [("Banded", _records(4)), ("Plain", _records(4))],
        str(path),
        banded_rows={"Banded": BAND},
    )

    wb = openpyxl.load_workbook(path)
    assert _fills(wb["Banded"], 3, 5) == [BAND_ARGB] * 5
    assert _fills(wb["Plain"], 3, 5) == [NO_FILL] * 5


def test_multi_sheet_general_key_applies_to_all(tmp_path):
    path = tmp_path / "general.xlsx"
    write_worksheets(
        [("A", _records(4)), ("B", _records(4))],
        str(path),
        banded_rows={"general": BAND},
    )
    wb = openpyxl.load_workbook(path)
    for name in ("A", "B"):
        assert _fills(wb[name], 3, 5) == [BAND_ARGB] * 5


def test_layout_defaults_unchanged(tmp_path):
    """No layout arguments must leave output exactly as before."""
    plain = tmp_path / "plain.xlsx"
    write_worksheet(_records(4), str(plain))
    ws = openpyxl.load_workbook(plain).active
    assert [c.value for c in ws[1]] == ["name", "score", "count", "ok", "when"]
    assert _fills(ws, 2, 5) == [NO_FILL] * 5
    assert not ws.merged_cells.ranges


# --- DataFrame paths ------------------------------------------------------
# Records, Arrow, Pandas and Polars each write cells through a different code
# path, so banding has to be asserted on all of them, not just Records.


@pytest.mark.parametrize("frame", ["pandas", "polars"])
def test_banding_on_dataframe_paths(tmp_path, frame):
    mod = pytest.importorskip(frame)
    data = {
        "name": ["a", "b", "c", "d"],
        "score": [1.5, 2.5, 3.5, 4.5],
        "count": [1, 2, 3, 4],
    }
    df = mod.DataFrame(data)

    path = tmp_path / f"{frame}.xlsx"
    write_worksheet(df, str(path), float_format="0.00", banded_rows=BAND)

    ws = openpyxl.load_workbook(path).active
    assert [c.value for c in ws[1]] == ["name", "score", "count"]
    assert _fills(ws, 2, 3) == [NO_FILL] * 3
    assert _fills(ws, 3, 3) == [BAND_ARGB] * 3, f"{frame}: hole on banded row"
    assert _fills(ws, 5, 3) == [BAND_ARGB] * 3
    assert ws.cell(row=3, column=2).value == pytest.approx(2.5)


@pytest.mark.parametrize("frame", ["pandas", "polars"])
def test_header_row_offset_on_dataframe_paths(tmp_path, frame):
    mod = pytest.importorskip(frame)
    df = mod.DataFrame({"a": [1, 2], "b": [3, 4]})

    path = tmp_path / f"{frame}-offset.xlsx"
    write_worksheet(
        df,
        str(path),
        header_row=1,
        merge_ranges=[(0, 0, 0, 1, "Banner")],
    )

    ws = openpyxl.load_workbook(path).active
    assert ws["A1"].value == "Banner"
    assert [c.value for c in ws[2]] == ["a", "b"]
    assert [c.value for c in ws[3]] == [1, 3]


def test_banding_on_datetime_dataframe_column(tmp_path):
    """Datetimes rely on a column format, which cannot alternate per row."""
    pd = pytest.importorskip("pandas")
    df = pd.DataFrame(
        {
            "when": pd.to_datetime(
                ["2026-01-01", "2026-01-02", "2026-01-03", "2026-01-04"]
            ),
            "n": [1, 2, 3, 4],
        }
    )
    path = tmp_path / "dt.xlsx"
    write_worksheet(df, str(path), banded_rows=BAND)

    ws = openpyxl.load_workbook(path).active
    assert _fills(ws, 3, 2) == [BAND_ARGB] * 2
    assert _fills(ws, 5, 2) == [BAND_ARGB] * 2
    assert ws.cell(row=3, column=1).value == datetime.datetime(2026, 1, 2)
    # The datetime number format must survive the banding.
    assert "yyyy" in ws.cell(row=3, column=1).number_format


class _FakeFrame:
    """Pandas-shaped object without ``__arrow_c_stream__``.

    Modern pandas and Polars both expose the Arrow stream, so the
    ``write_dataframe`` fallback is unreachable through them. This stands in
    for the old-pandas / exotic-dtype case that path actually exists for.
    """

    class _Dtype:
        def __init__(self, kind):
            self.kind = kind

    class _Series:
        def __init__(self, values):
            self._values = values

        def tolist(self):
            return list(self._values)

    def __init__(self, data, kinds):
        self._data = data
        self.columns = list(data)
        self.dtypes = [self._Dtype(k) for k in kinds]

    def __len__(self):
        return len(next(iter(self._data.values())))

    def __getitem__(self, key):
        return self._Series(self._data[key])


def test_banding_on_dataframe_fallback_path(tmp_path):
    df = _FakeFrame(
        {
            "name": ["a", "b", "c", "d"],
            "score": [1.5, 2.5, 3.5, 4.5],
            "count": [1, 2, 3, 4],
            "ok": [True, False, True, False],
            "when": [datetime.datetime(2026, 1, d + 1) for d in range(4)],
        },
        kinds=["O", "f", "i", "b", "M"],
    )

    path = tmp_path / "fallback.xlsx"
    write_worksheet(df, str(path), float_format="0.00", banded_rows=BAND)

    ws = openpyxl.load_workbook(path).active
    assert [c.value for c in ws[1]] == ["name", "score", "count", "ok", "when"]
    assert _fills(ws, 2, 5) == [NO_FILL] * 5
    assert _fills(ws, 3, 5) == [BAND_ARGB] * 5
    assert _fills(ws, 5, 5) == [BAND_ARGB] * 5
    assert ws.cell(row=3, column=1).value == "b"
    assert ws.cell(row=3, column=2).value == pytest.approx(2.5)
    assert ws.cell(row=3, column=4).value is False
    assert ws.cell(row=3, column=5).value == datetime.datetime(2026, 1, 2)


def test_header_row_offset_on_dataframe_fallback_path(tmp_path):
    df = _FakeFrame({"a": [1, 2], "b": [3, 4]}, kinds=["i", "i"])
    path = tmp_path / "fallback-offset.xlsx"
    write_worksheet(
        df, str(path), header_row=1, merge_ranges=[(0, 0, 0, 1, "Banner")]
    )
    ws = openpyxl.load_workbook(path).active
    assert ws["A1"].value == "Banner"
    assert [c.value for c in ws[2]] == ["a", "b"]
    assert [c.value for c in ws[3]] == [1, 3]
