import openpyxl

from rustpy_xlsxwriter import FastExcel


class TestFreezePanesSingleSheet:
    def test_freeze_row(self, tmp_path, small_records):
        path = str(tmp_path / "freeze_row.xlsx")
        FastExcel(path).freeze(row=2).sheet("Sheet1", small_records).save()
        wb = openpyxl.load_workbook(path)
        # set_freeze_panes(2, 0) -> openpyxl cell ref "A3" (row+1 in 1-indexed)
        assert wb.active.freeze_panes == "A3"
        wb.close()

    def test_freeze_col(self, tmp_path, small_records):
        path = str(tmp_path / "freeze_col.xlsx")
        FastExcel(path).freeze(col=1).sheet("Sheet1", small_records).save()
        wb = openpyxl.load_workbook(path)
        assert wb.active.freeze_panes == "B1"
        wb.close()

    def test_freeze_both(self, tmp_path, small_records):
        path = str(tmp_path / "freeze_both.xlsx")
        FastExcel(path).freeze(row=2, col=1).sheet("Sheet1", small_records).save()
        wb = openpyxl.load_workbook(path)
        assert wb.active.freeze_panes == "B3"
        wb.close()

    def test_no_freeze(self, tmp_path, small_records):
        path = str(tmp_path / "no_freeze.xlsx")
        FastExcel(path).sheet("Sheet1", small_records).save()
        wb = openpyxl.load_workbook(path)
        assert wb.active.freeze_panes is None
        wb.close()


class TestFreezePanesMultiSheet:
    def test_general_freeze_applies_to_all(self, tmp_path, small_records):
        path = str(tmp_path / "multi_freeze.xlsx")
        (
            FastExcel(path)
            .freeze(row=1)
            .sheet("Sheet1", small_records)
            .sheet("Sheet2", small_records)
            .save()
        )
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"].freeze_panes == "A2"
        assert wb["Sheet2"].freeze_panes == "A2"
        wb.close()

    def test_per_sheet_freeze(self, tmp_path, small_records):
        path = str(tmp_path / "multi_freeze_specific.xlsx")
        (
            FastExcel(path)
            .freeze(row=1, sheet="Sheet1")
            .freeze(row=3, col=2, sheet="Sheet2")
            .sheet("Sheet1", small_records)
            .sheet("Sheet2", small_records)
            .save()
        )
        wb = openpyxl.load_workbook(path)
        assert wb["Sheet1"].freeze_panes == "A2"
        assert wb["Sheet2"].freeze_panes == "C4"
        wb.close()
